# Copyright (c) 2023 StrutNut
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense
# copies of the Software, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

# Notwithstanding the foregoing, you MAY NOT use the Software for commercial purposes.

# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, CAUSED BY THE SOFTWARE. THE ENTIRE RISK AS TO THE
# QUALITY AND PERFORMANCE OF THE SOFTWARE IS WITH YOU. SHOULD THE SOFTWARE PROVE
# DEFECTIVE, YOU ASSUME THE COST OF ALL NECESSARY SERVICING, REPAIR OR CORRECTION.

import os
import requests
import re
from lxml import etree
import argparse
import openpyxl

# 初始化
ver = "1.0.0"
xname = "/html/body/div/div[2]/div/div/div[1]/div[2]/text()"
xnum = "/html/body/div/div[2]/div/div/div[5]/div[2]/text()"
xdomain = "/html/body/div/div[2]/div/div/div[2]/div[2]/text()"
xdesc = "/html/body/div/div[2]/div/div/div[4]/div[2]/text()"
xhome = "/html/body/div/div[2]/div/div/div[3]/div[2]/a/@href"
xowner = "/html/body/div/div[2]/div/div/div[6]/div[2]/text()"
xuptime = "/html/body/div/div[2]/div/div/div[7]/div[2]/text()"
xstatus = "/html/body/div/div[2]/div/div/div[8]/div[2]/text()"
xpic = "/html/head/meta[11]/@content"
not_exist = '/html/body/div/div[2]/div/form/h4[1]'
recycle = '/html/body/div/div[2]/div/div/p[5]'


def init():
    parser = argparse.ArgumentParser()

    parser.add_argument('--start', type=int, default=20220000)
    parser.add_argument('--end', type=int, default=20220000)
    parser.add_argument('--output', type=str, default='./萌备案数据.xlsx')
    parser.add_argument('--append', action='store_true')

    args = parser.parse_args()

    start = args.start
    end = args.end

    title_list = ['网站名称', '网站域名', '网站首页', '网站信息', '萌备案号', '所有者', '更新时间', '状态', '图片链接']
    data_dict = {title: [] for title in title_list}

    num = start
    reset_time = 0

    while num <= end:

        url = "https://icp.gov.moe/?keyword=" + str(num)
        headers = {
            'User-Agent': 'MoeICP-CLI/' + ver
        }
        resp = requests.get(url, headers=headers)
        resp.encoding = 'utf-8'

        html = etree.HTML(resp.text)

        if len(html.xpath(not_exist)) > 0:
            print('萌号 ' + str(num) + ' 未有记录！')
            num = num + 1
            continue
        if len(html.xpath(recycle)) > 0:
            print('萌号 ' + str(num) + ' 已被回收！')
            num = num + 1
            continue

        res = get_info(html, data_dict)
        if not res:
            print('萌号：' + str(num) + ' 抓取失败,重试第 ' + str(reset_time + 1) + ' 次。')
            reset_time = reset_time + 1
            if reset_time == 10:
                print('萌号：' + str(num) + ' 的抓取尝试全都失败了，进入下一个吧...')
                reset_time = 0
                num = num + 1
            continue

        reset_time = 0
        print('萌号 ' + str(num) + ' 抓取成功！')
        num = num + 1

    last_row = 1
    file_path = args.output
    if not args.append and os.path.exists(file_path):
        os.remove(file_path)

    wb = openpyxl.Workbook()
    sheet = wb.active
    if args.append:
        if os.path.exists(file_path):
            wbk = openpyxl.load_workbook(file_path)
        else:
            wbk = openpyxl.Workbook()
        sheetk = wbk.active
        last_row = sheetk.max_row
        for row in range(1, sheetk.max_row + 1):
            for col in range(1, sheetk.max_column + 1):
                sheet.cell(row=row, column=col).value = sheetk.cell(row=row, column=col).value

    output_info(data_dict, sheet, last_row)
    wb.save(file_path)
    absolute_path = os.path.abspath(file_path)
    print('任务完成，目标文件保存在 ' + absolute_path + ' 中。')


def get_info(html, data_dict):
    try:
        xname_data = html.xpath(xname)[0]
    except:
        return False

    compileint = re.compile('\d+')
    inttext = compileint.findall(html.xpath(xnum)[0])
    xnumber_data = inttext[0]
    xdomain_data = html.xpath(xdomain)[0]
    xdescription_data = html.xpath(xdesc)[0]
    xhome_data = html.xpath(xhome)[0]
    xowner_data = html.xpath(xowner)[0]
    xuptime_data = html.xpath(xuptime)[0]
    xstatus_data = html.xpath(xstatus)[0]
    xpiclink_data = html.xpath(xpic)[0]

    msg_list = [xname_data, xdomain_data, xhome_data, xdescription_data, '萌ICP备' + xnumber_data + '号'
        , xowner_data, xuptime_data, xstatus_data, xpiclink_data]
    for i, title in enumerate(data_dict):
        data_dict[title].append(msg_list[i])

    return True


def output_info(data_dict, sheet, last_row):
    for col, title in enumerate(data_dict, 1):
        sheet.cell(row=1, column=col).value = title

    for i in range(len(data_dict['网站名称'])):
        for j, (title, data) in enumerate(data_dict.items(), 1):
            sheet.cell(row=last_row + i + 1, column=j).value = data[i]


if __name__ == "__main__":
    init()
